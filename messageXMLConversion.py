#Import modules
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from tkinter import scrolledtext
from tkinter import messagebox
import xml.etree.ElementTree as ET
import logging
import getpass
import os
import sys
import subprocess
from pathlib import Path

#Define functions
def needExport():
    response = input("Do you need to export the Excel files? [Y/N]: ")
    #If exports are needed get login credentials
    if response in questionYes:
        logger.info("You chose to export the Excel files.")

        #Prompt to select generate script
        selectGenerateFile()
    elif response in questionNo:
        logger.info("You chose to skip exporting the Excel files.")
        generateFile = 0
        getXML(generateFile)
    else:
        print("Please enter 'yes' or 'no'...")
        needExport()

def showErrorMessage(fileName, errorList):
    windowTitle = fileName + " Errors: Continue?"
    topLabel = "The Excel files for the " + fileName + " had the following errors. The Excel files will be skipped from being updated.\n\n"
    bottomLabel = "\n\nDo you want to continue with the rest of the selected CSC XML files?\n\n"
    errorWindow = tk.Toplevel()
    errorWindow.title(windowTitle)

    #Frame for error message
    errorFrame = ttk.Frame(errorWindow)
    errorFrame.grid(row=0, column=0, columnspan=3, ipadx=10, ipady=10, sticky="N, W, E, S")
    errorWindow.columnconfigure(0, weight=1)
    errorWindow.rowconfigure(1, weight=1)
    errorFrame.columnconfigure(0, weight=1)
    errorFrame.rowconfigure(0, weight=1)

    beforeErrors = ttk.Label(errorFrame, text=topLabel).grid(row=0, column=0, columnspan=3)
    errorArea = scrolledtext.ScrolledText(errorFrame, wrap="word", width=75, height=10)
    errorArea.grid(row=1, column=0, sticky="N, W, E, S")
    errorArea.insert(tk.INSERT, errorList)
    afterErrors = ttk.Label(errorFrame, text=bottomLabel).grid(row=2, column=0, columnspan=3)

    #Button frame
    buttonFrame = ttk.Frame(errorWindow)
    buttonFrame.grid(row=1, column=0, columnspan=2, ipadx=10, ipady=10, sticky="W, E")
    buttonFrame.columnconfigure(0, weight=1)
    buttonFrame.rowconfigure(1, weight=1)
    cancelButton = ttk.Button(buttonFrame, text="Cancel", command=quitProcess).grid(row=1, column=0, padx=10, pady=10, sticky="SE")
    okButton = ttk.Button(buttonFrame, text="OK", default="active", command= lambda: continueProcess(fileName, errorWindow)).grid(row=1, column=1, padx=10, pady=10, sticky="SE")

    #Center the window
    myLeftPos = (errorWindow.winfo_screenwidth() - 850) / 2
    myTopPos = (errorWindow.winfo_screenheight() - 350) / 2
    errorWindow.geometry( "%dx%d+%d+%d" % (850, 350, myLeftPos, myTopPos))

    #Require a button click
    errorWindow.protocol("WM_DELETE_WINDOW", disableEvent)
    errorArea.configure(state ='disabled')
    errorWindow.resizable(False, False)

    #Open the window
    errorWindow.mainloop()

def continueProcess(fileName, errorWindow):
    logger.info("You chose to continue the proccessing the files. You need to rerun this script for {%s}." % fileName)
    errorWindow.quit()
    errorWindow.destroy()

def quitProcess():
    logger.info("You chose to quit the proccessing the files.")
    sys.exit()

def disableEvent():
    pass

def selectGenerateFile():
    root.scriptFilename = filedialog.askopenfilename(initialdir=home, title='Please select the MagicDraw generate.bat (Windows) or generate.sh (macOS) file', filetypes=[("generate.bat", "*.bat"), ("generate.sh", "*.sh")])
    generateFile = root.scriptFilename
    path, filename = os.path.split(generateFile)
    if filename == "generate.sh" or filename == "generate.bat":
        logger.info("You chose: %s" % generateFile)
        getXML(generateFile)
    else:
        logger.warning("You must select either a generate.bat or generate.sh file. This should be located at <yourMagicDrawInstallLocation>/plugins/com.nomagic.magicdraw.reportwizard/.")
        selectGenerateFile()
def getXML(generateFile):
    origXML = currdir + "/XML/sal_interfaces"
    root.xmlFilename = filedialog.askopenfilenames(initialdir=origXML, title='Please select the GitHub XML file(s)', filetypes=[("XML Files", "*.xml")])
    selectedFiles = root.xmlFilename
    xmlFileList = list(selectedFiles)
    logger.info("You chose: %s" % xmlFileList)
    #If we need to export the Excel Files
    if len(xmlFileList) > 0:
        exportResult = 0
        if generateFile != 0:
            exportResult = exportExcel(xmlFileList, generateFile)
        if exportResult == 0 :
            for xmlFileLocation in xmlFileList:
                #Temporarily set path to same directory as python file
                path, filename = os.path.split(xmlFileLocation)
                path = currdir
                fileName = filename
                fileBase = filename.split('.')[0]
                messageType = fileBase.split('_')[1]
                cscName = fileBase.split('_')[0]
                logger.info("Preparing file: %s" % xmlFileLocation)
                preparedFile = prepareExcelFiles(path, fileName, fileBase, messageType, cscName, xmlFileLocation, generateFile)
                if preparedFile == 1:
                    continue
        else:
            messagebox.showerror("Excel Export Errors", exportResult)
            #sys.exit()
    else:
        logger.warning("You must select an XML file.")
        getXML(generateFile)
def exportExcel(xmlFileList, generateFile):
    propArray = []
    for xmlFileLocation in xmlFileList:
        #Temporarily set path to same directory as python file
        path, filename = os.path.split(xmlFileLocation)
        path = currdir
        fileName = filename
        fileBase = filename.split('.')[0]
        messageType = fileBase.split('_')[1]
        cscName = fileBase.split('_')[0]
        #Set the file locations
        fileLocation = path + "/" + cscName + "/"
        excelFileLocation = fileLocation + fileBase + ".xlsx"
        propFileLocation = fileLocation + fileBase + ".properties"

        #Create the property file
        project = 'project = Telescope & Site Software Components\n'
        package = 'package = ' + cscName + '::Signals::' + messageType + '\n'
        template = 'template = CSC Signal Export\n'
        output = 'output = ' + excelFileLocation

        propFile = open(propFileLocation, "w")
        propFile.write(project)
        propFile.write(package)
        propFile.write(template)
        propFile.write(output)
        propFile.close()
        propArray.append(propFileLocation)

    propList = '" "'.join(propArray)
    #Set the property argument
    propArg = '-properties "' + propList + '"'

    exportResult = subprocess.run([generateFile, '-server "twcloud.lsst.org"', '-servertype "twcloud"', '-ssl true', '-login "rgenerator"', '-spassword "4e0e16bb44e1c62fe007b776d088c899c585425cbf6f424161867c13735f7f67e683f676e4eb811e4ad464e04c822cb14c1c8a385bc7461b8e94ae7292a31ee9e2758a20124d77b7c79546c6be4878db10f021f91f147833cc63fb93b8aefb6f13110dd959128c49a1da292c7ee8f98640982e1e75b40abc397514d4712d471c"', '-leaveprojectopen true', propArg], check=True)
    result = exportResult.returncode
    return result

def prepareExcelFiles(path, fileName, fileBase, messageType, cscName, xmlFileLocation, generateFile):
    errorArray = []
    errorFlag = 0
    excelFileLocation = path + "/" + cscName + "/" + fileBase + ".xlsx"

    #Verify the columns
    #Get message Excel file
    if os.path.isfile(excelFileLocation):
        wb = load_workbook(filename = excelFileLocation)
        #Verify message Excel ws is in correct format
        if 'Signals' in wb.sheetnames:
            ws = wb['Signals']
            if (ws['A1'].value == 'Name') and (ws['B1'].value == 'Alias') and (ws['C1'].value == 'Documentation') and (ws['D1'].value == 'Subsystem') and (ws['E1'].value == 'Version') and (ws['F1'].value == 'Author') and (ws['G1'].value == 'Device') and (ws['H1'].value == 'Property') and (ws['I1'].value == 'Action') and (ws['J1'].value == 'Value') and (ws['K1'].value == 'Explanation') and (ws['L1'].value == 'Order') and (ws['M1'].value == 'Element ID'):
                logger.info("The {%s} file has the required format." % excelFileLocation)

                #Clearing the SyncAction Column in Signals
                logger.info("Clearing the 'Signals' sheet in the {%s} SyncAction column." % excelFileLocation)
                ws.delete_cols(14)
                ws.insert_cols(14)
                ws.cell(row=1, column =14, value = 'SyncAction')

                #Verify message parameter Excel ws is in correct format
                if 'Signal Properties' in wb.sheetnames:
                    ws_p = wb['Signal Properties']
                    if (ws_p['A1'].value == 'Owner') and (ws_p['B1'].value == 'Name') and (ws_p['C1'].value == 'Documentation') and (ws_p['D1'].value == 'Type') and (ws_p['E1'].value == 'Size') and (ws_p['F1'].value == 'Type Modifier') and (ws_p['G1'].value == 'Multiplicity') and (ws_p['H1'].value == 'Order') and (ws_p['I1'].value == 'Element ID'):
                        logger.info("The 'Signal Properties' sheet in the {%s} file has the required format." % excelFileLocation)

                        #Clearing the SyncAction Column in signal properties
                        logger.info("Clearing the {%s} Signal Properties sheet SyncAction column." % excelFileLocation)
                        ws_p.delete_cols(10)
                        ws_p.insert_cols(10)
                        ws_p.cell(row=1, column =10, value = 'SyncAction')

                        #Verify enumeration Excel ws is in correct format
                        if 'Enumerations' in wb.sheetnames:
                            ws_e = wb['Enumerations']
                            if (ws_e['A1'].value == 'Name') and (ws_e['B1'].value == 'Order') and (ws_e['C1'].value == 'Element ID'):
                                logger.info("The Enumerations sheet in the {%s} file has the required format." % excelFileLocation)

                                #Clearing the SyncAction Column in enumerations
                                logger.info("Clearing the {%s} Enumerations sheet SyncAction column." % excelFileLocation)
                                ws_e.delete_cols(4)
                                ws_e.insert_cols(4)
                                ws_e.cell(row=1, column =4, value = 'SyncAction')

                                #Verify enumeration literal Excel ws is in correct format
                                if 'Enumeration Options' in wb.sheetnames:
                                    ws_l = wb['Enumeration Options']
                                    if (ws_l['A1'].value == 'Owner') and (ws_l['B1'].value == 'Name') and (ws_l['C1'].value == 'Order') and (ws_l['D1'].value == 'Element ID'):
                                        logger.info("The 'Enumeration Options' sheet in the {%s} file has the required format." % excelFileLocation)

                                        #Clearing the SyncAction Column in enemuration literals
                                        logger.info("Clearing the {%s} SyncAction column." % excelFileLocation)
                                        ws_l.delete_cols(5)
                                        ws_l.insert_cols(5)
                                        ws_l.cell(row=1, column =5, value = 'SyncAction')

                                    else:
                                        logger.error("The 'Enumeration Options' sheet in the {%s} file columns do not follow the required format of [Owner, Name, Order]. Please fix the format and run again." % excelFileLocation)
                                        errorFlag = 1
                                        errorMessage = "The 'Enumeration Options' sheet in the {" + excelFileLocation + "} file columns do not follow the required format. The Excel file will NOT be updated."
                                        errorArray.append(errorMessage)
                                else:
                                    logger.error("Could not find 'Enumeration Options' sheet in {%s} file." % excelFileLocation)
                                    errorFlag = 1
                                    errorMessage = "Could not find 'Enumeration Options' sheet in the {" + excelFileLocation + "} file. The Excel file will NOT be updated."
                                    errorArray.append(errorMessage)

                            else:
                                logger.error("The Enumerations sheet in the {%s} file columns do not follow the required format of [Name, Order]. Please fix the format and run again." % excelFileLocation)
                                errorFlag = 1
                                errorMessage = "The Enumerations sheet in the {" + excelFileLocation + "} file columns do not follow the required format. The Excel file will NOT be updated."
                                errorArray.append(errorMessage)
                        else:
                            logger.error("Could not find 'Enumerations' sheet in the {%s} file." % excelFileLocation)
                            errorFlag = 1
                            errorMessage = "Could not find 'Enumerations' sheet in the {" + excelFileLocation + "} file. The Excel file will NOT be updated."
                            errorArray.append(errorMessage)

                    else:
                        logger.error("The 'Signal Properties' sheet in the {%s} file columns do not follow the required format of [Owner, Name, Documentation, Type, Size, Type Modifier, Multiplicity, Order]. Please fix the format and run again." % excelFileLocation)
                        errorFlag = 1
                        errorMessage = "The 'Signal Properties' sheet in the {" + excelFileLocation + "} file columns do not follow the required format. The Excel file will NOT be updated."
                        errorArray.append(errorMessage)
                else:
                    logger.error("Could not find 'Signal Properties' sheet in {%s} file." % excelFileLocation)
                    errorFlag = 1
                    errorMessage = "Could not find 'Signal Properties' sheet in {" + excelFileLocation + "} file. The Excel file will NOT be updated."
                    errorArray.append(errorMessage)

            else:
                logger.error("The 'Signals' sheet in the {%s} file columns do not follow the required format of [Name, Alias, Documentation, Subsystem, Version, Author, Device, Property, Action, Value, Explanation, Order]. Please fix the format and run again." % excelFileLocation)
                errorFlag = 1
                errorMessage = "The 'Signals' sheet in the {" + excelFileLocation + "} file columns do not follow the required format. The Excel file will NOT be updated."
                errorArray.append(errorMessage)
        else:
            logger.error("Could not find 'Signals' sheet in the {%s} file." % excelFileLocation)
            errorFlag = 1
            errorMessage = "Could not find 'Signals' sheet in the {" + excelFileLocation + "} file. The Excel file will NOT be updated."
            errorArray.append(errorMessage)
    else:
        logger.error("Could not find the Excel file at the following location: %s." % excelFileLocation)
        errorFlag = 1
        errorMessage = "Could not find the Excel file at the following location {" + excelFileLocation + "}. The Excel file will NOT be updated."
        errorArray.append(errorMessage)

    #Check if no errors continue to updating the Excel files
    if errorFlag == 0:
        updatedFiles = updateExcelFiles(xmlFileLocation, excelFileLocation, wb, ws, ws_p, ws_e, ws_l, messageType, fileName)
    else:
        errorList = "\n".join(errorArray)
        showErrorMessage(fileName, errorList)


def updateExcelFiles(xmlFileLocation, excelFileLocation, wb, ws, ws_p, ws_e, ws_l, messageType, fileName):
    tree = ET.parse(xmlFileLocation)
    treeRoot = tree.getroot()

    salMessages = []
    salAlias = []
    messageParams = []
    messageTypeEnums = []
    enumLiterals = []
    updateErrorArray = []
    updateErrorFlag = 0
    messageOrder = 0
    messagePropOrder = 0
    enumOrder = 0
    enumLitOrder = 0

    #Check for the message type
    if messageType == "Commands":
        messageRoot = "SALCommand"
    elif messageType == "Events":
        messageRoot = "SALEvent"
    elif messageType == "Telemetry":
        messageRoot = "SALTelemetry"
    else:
        logger.error("Unknown message type of %s." % messageType)
        updateErrorFlag = 1
        updateErrorMessage = "Unknown message type of {" + messageType + "}. The Excel files for {" + fileName +"} will NOT be updated."
        updateErrorArray.append(updateErrorMessage)

    if updateErrorFlag == 0:
        #Check for enumerations
        logger.info("Checking the {%s} file for enumerations and enumeration options." % fileName)
        for enum in treeRoot.findall('Enumeration'):
            enumOrder = enumOrder + 1
            count_e = 0
            enumNm = enum.text.split('_')[0]
            enumName = enumNm.lstrip()
            enumNameLower = enumName[0].lower() + enumName[1:]
            messageTypeEnums.append(enumNameLower)

            #Check if existing enumeration row is found
            for row in ws_e.iter_rows(min_row=2, max_col=1):
                for cell in row:
                    if cell.value == enumNameLower:
                        count_e = count_e + 1
                        thisRow_e = cell.row
                        ws_e.cell(row=thisRow_e, column=2, value=str(enumOrder))
                        ws_e.cell(row=thisRow_e, column=4, value="Update")

            #If no existing enumeration row found, create a new row
            if count_e == 0:
                filteredNumRows_e = list(filter(None, ws_e['A']))
                totalRows_e = len(filteredNumRows_e)
                newRow_e = totalRows_e + 1
                ws_e.cell(row=newRow_e, column=1, value=enumNameLower)
                ws_e.cell(row=newRow_e, column=2, value=str(enumOrder))
                ws_e.cell(row=newRow_e, column=4, value="Add")
            elif count_e > 1:
                logger.error("There are duplicate rows for enumeration {%s}" % enumNameLower)

            #Now start on literals
            enumLits = enum.text.replace(" ", "").split(',')
            for lit in enumLits:
                enumLitOrder = enumLitOrder + 1
                count_l = 0
                litOwn = lit.split('_')[0]
                litOwner = litOwn.lstrip()
                litOwnerLower = litOwner[0].lower() + litOwner[1:]
                lst = lit.lstrip()
                litName = lst.split('_',1)[1]
                litTup = (litOwnerLower, litName)
                enumLiterals.append(litTup)

                #Update literal row if existing literal is found
                for row in ws_l.iter_rows(min_row=2, max_col=2):
                    rowList = []
                    for cell in row:
                        rowList.append(cell.value)
                        thisRow = cell.row
                    tupRow = tuple(rowList)

                    if tupRow == litTup:
                        count_l = count_l + 1
                        ws_l.cell(row=thisRow, column=3, value=str(enumLitOrder))
                        ws_l.cell(row=thisRow, column=5, value="Update")

                #If no existing literal row found, create a new row
                if count_l == 0:
                    filteredNumRows_l = list(filter(None, ws_l['A']))
                    totalRows_l = len(filteredNumRows_l)
                    newRow_l = totalRows_l + 1
                    ws_l.cell(row=newRow_l, column=1, value=litOwnerLower)
                    ws_l.cell(row=newRow_l, column=2, value=litName)
                    ws_l.cell(row=newRow_l, column=3, value=str(enumLitOrder))
                    ws_l.cell(row=newRow_l, column=5, value="Add")
                elif count_l > 1:
                    errorMessage = litTup
                    logger.error("There are duplicate rows for enumeration option {%s}" % (errorMessage,))
                del litTup

        #Remove any enumeration rows not in XML
        i=2
        while i <= ws_e.max_row:
            cellVal = ws_e.cell(row=i, column=1).value
            if cellVal == None or cellVal == "":
                ws_e.delete_rows(i)
                continue
            elif (cellVal not in messageTypeEnums) :
                ws_e.cell(row=i, column=4, value="Delete")
                logger.info("The row for enumeration with the name {%s} is not found in the XML and has been marked for deletion." % cellVal)
            i += 1

        #Remove any literal rows not in XML
        i=2
        while i <= ws_l.max_row:
            tupVal = (ws_l.cell(row=i, column=1).value, ws_l.cell(row=i, column=2).value)
            if tupVal == (None, None) or tupVal == ("", "") or tupVal == (None, "") or tupVal == ("", None):
                ws_l.delete_rows(i)
                continue
            elif (tupVal not in enumLiterals) :
                ws_l.cell(row=i, column=5, value="Delete")
                logger.info("The row for enumeration literal with the name {%s} is not found in the XML and has been marked for deletion." % (tupVal,))
            i += 1


        #Check for messages
        logger.info("Checking the {%s} file for messages and message parameters." % fileName)
        for member in treeRoot.findall(messageRoot):
            messageOrder = messageOrder + 1
            count = 0
            alias = ""
            topic = ""
            subsystem = ""
            version = ""
            author = ""
            explanation = ""
            items = ""
            device = ""
            prop = ""
            action = ""
            value = ""
            desc = ""

            #Get values
            if member.find('EFDB_Topic') != None:
                topic = member.find('EFDB_Topic').text
                if topic == None or topic == "":
                    logger.error("The message at order number {%s} has no EFDB_Topic." % str(messageOrder))
                    updateErrorFlag = 1
                    updateErrorMessage = "The message at order number {" + str(messageOrder) + "} has no EFDB_Topic. The Excel files for {" + fileName +"} will NOT be updated."
                    updateErrorArray.append(updateErrorMessage)
            else:
                logger.error("The message at order number {%s} has no EFDB_Topic tag." % str(messageOrder))
                updateErrorFlag = 1
                updateErrorMessage = "The message at order number {" + str(messageOrder) + "} has no EFDB_Topic tag. The Excel files for {" + fileName +"} will NOT be updated."
                updateErrorArray.append(updateErrorMessage)
            if member.find('Alias') != None:
                alias = member.find('Alias').text
            if member.find('Subsystem') != None:
                subsystem = member.find('Subsystem').text
            if member.find('Version') != None:
                version = member.find('Version').text
            if member.find('Author') != None:
                author = member.find('Author').text
            if member.find('Explanation') != None:
                explanation = member.find('Explanation').text
            if member.find('item') != None:
                items = member.findall('item')
            if member.find('Device') != None:
                device = member.find('Device').text
            if member.find('Property') != None:
                prop = member.find('Property').text
            if member.find('Action') != None:
                action = member.find('Action').text
            if member.find('Value') != None:
                value = member.find('Value').text
            if member.find('Description') != None:
                desc = member.find('Description').text

            #Add message topic and alias to lists
            salMessages.append(topic)
            salAlias.append(alias)


            #Update message row if existing message is found
            for row in ws.iter_rows(min_row=2, max_col=1):
                for cell in row:
                    if cell.value == topic:
                        count += 1
                        thisRow = cell.row
                        ws.cell(row=thisRow, column=2, value=alias)
                        ws.cell(row=thisRow, column=3, value=desc)
                        ws.cell(row=thisRow, column=4, value=subsystem)
                        ws.cell(row=thisRow, column=5, value=version)
                        ws.cell(row=thisRow, column=6, value=author)
                        ws.cell(row=thisRow, column=7, value=device)
                        ws.cell(row=thisRow, column=8, value=prop)
                        ws.cell(row=thisRow, column=9, value=action)
                        ws.cell(row=thisRow, column=10, value=value)
                        ws.cell(row=thisRow, column=11, value=explanation)
                        ws.cell(row=thisRow, column=12, value=str(messageOrder))
                        ws.cell(row=thisRow, column=14, value="Update")

            #If no existing message row found, create a new row
            if count == 0:
                filteredNumRows = list(filter(None, ws['A']))
                totalRows = len(filteredNumRows)
                newRow = totalRows + 1
                ws.cell(row=newRow, column=1, value=topic)
                ws.cell(row=newRow, column=2, value=alias)
                ws.cell(row=newRow, column=3, value=desc)
                ws.cell(row=newRow, column=4, value=subsystem)
                ws.cell(row=newRow, column=5, value=version)
                ws.cell(row=newRow, column=6, value=author)
                ws.cell(row=newRow, column=7, value=device)
                ws.cell(row=newRow, column=8, value=prop)
                ws.cell(row=newRow, column=9, value=action)
                ws.cell(row=newRow, column=10, value=value)
                ws.cell(row=newRow, column=11, value=explanation)
                ws.cell(row=newRow, column=12, value=(messageOrder))
                ws.cell(row=newRow, column=14, value="Add")
            elif count > 1:
                logger.error("There are duplicate rows for message {%s}" % alias)

            #Now start on the parameters
            for item in items:
                messagePropOrder = messagePropOrder + 1
                count_p = 0
                param_name = ""
                param_desc = ""
                idl_type = ""
                idl_size = ""
                unit = ""
                param_count = ""


                #Get values
                if item.find('EFDB_Name') != None:
                    param_name = item.find('EFDB_Name').text
                else:
                    logger.error("The message parameter at order number {%s} has no EFDB_Name." % messagePropOrder)
                    updateErrorFlag = 1
                    updateErrorMessage = "The message parameter at order number {" + messagePropOrder + "} has no EFDB_Name. The Excel files for {" + fileName +"} will NOT be updated."
                    updateErrorArray.append(updateErrorMessage)
                if item.find('Description') != None:
                    param_desc = item.find('Description').text
                if item.find('IDL_Type') != None:
                    if alias in messageTypeEnums:
                        idl_type = alias
                    elif param_name in messageTypeEnums:
                        idl_type = param_name
                    else:
                        idl_type = item.find('IDL_Type').text
                        if idl_type == "string":
                            idl_type = idl_type[0].upper() + idl_type[1:]
                if item.find('IDL_Size') != None:
                    idl_size = int(item.find('IDL_Size').text)
                if item.find('Units') != None:
                    unit = item.find('Units').text
                if item.find('Count') != None:
                    param_count = item.find('Count').text

                #Create tuple of owner and name
                tup = (topic, param_name)

                #Add param tuple to list
                messageParams.append(tup)

                #Update parameter row if existing message is found
                for row in ws_p.iter_rows(min_row=2, max_col=2):
                    rowList = []
                    for cell in row:
                        rowList.append(cell.value)
                        thisRow = cell.row
                    tupRow = tuple(rowList)

                    if tupRow == tup:
                        count_p += 1
                        ws_p.cell(row=thisRow, column=3, value=param_desc)
                        ws_p.cell(row=thisRow, column=4, value=idl_type)
                        ws_p.cell(row=thisRow, column=5, value=str(idl_size))
                        ws_p.cell(row=thisRow, column=6, value=unit)
                        ws_p.cell(row=thisRow, column=7, value=param_count)
                        ws_p.cell(row=thisRow, column=8, value=str(messagePropOrder))
                        ws_p.cell(row=thisRow, column=10, value="Update")

                #If no existing parameter row found, create a new row
                if count_p == 0:
                    filteredNumRows_p = list(filter(None, ws_p['A']))
                    totalRows_p = len(filteredNumRows_p)
                    newRow_p = totalRows_p + 1
                    ws_p.cell(row=newRow_p, column=1, value=topic)
                    ws_p.cell(row=newRow_p, column=2, value=param_name)
                    ws_p.cell(row=newRow_p, column=3, value=param_desc)
                    ws_p.cell(row=newRow_p, column=4, value=idl_type)
                    ws_p.cell(row=newRow_p, column=5, value=str(idl_size))
                    ws_p.cell(row=newRow_p, column=6, value=unit)
                    ws_p.cell(row=newRow_p, column=7, value=param_count)
                    ws_p.cell(row=newRow_p, column=8, value=str(messagePropOrder))
                    ws_p.cell(row=newRow_p, column=10, value="Add")

                elif count_p > 1:
                    errorMessage = tup
                    logger.error("There are duplicate rows for message parameters {%s}" % (errorMessage,))
                del tup

        #Remove any message rows not in XML
        i=2
        while i <= ws.max_row:
            cellVal = ws.cell(row=i, column=1).value
            if cellVal == None or cellVal == "":
               ws.delete_rows(i)
               continue
            elif (cellVal not in salMessages) :
                ws.cell(row=i, column=14, value="Delete")
                logger.info("The row for message with the name {%s} is not found in the XML and has been marked for deletion." % cellVal)
            i += 1


        #Remove any parameter rows not in XML
        i=2
        while i <= ws_p.max_row:
            tupVal = (ws_p.cell(row=i, column=1).value, ws_p.cell(row=i, column=2).value)
            if tupVal == (None, None) or tupVal == ("", "") or tupVal == (None, "") or tupVal == ("", None):
                ws_p.delete_rows(i)
                continue
            elif (tupVal not in messageParams) :
                ws_p.cell(row=i, column=10, value="Delete")
                logger.info("The row for message parameter with {%s} is not found in the XML and has been marked for deletion." % (tupVal,))
            i += 1


        #Check for unused enumerations
        logger.info("Checking the {%s} file for unused enumerations." % fileName)
        for e in messageTypeEnums:
            if e not in salAlias:
                logger.error("The enumeration {%s} is defined but not found as any message alias." % e)

        #If no errors save the Excel Files
        if updateErrorFlag == 0:
            #Save Excel files
            logger.info("Saving the {%s} file." % excelFileLocation)
            wb.save(excelFileLocation)
        else:
            #Show error message
            updateErrorList = "\n".join(updateErrorArray)
            showErrorMessage(fileName, updateErrorList)
    else:
        #Show error message
        updateErrorList = "\n".join(updateErrorArray)
        showErrorMessage(fileName, updateErrorList)

#Define global variables
root = tk.Tk()
root.withdraw() #use to hide tkinter window

currdir = os.getcwd()
home = str(Path.home())
xmlFileLocation = ""
excelFileLocation_Message = ""
excelFileLocation = ""
excelFileLocation = ""
excelFileLocation = ""
questionYes = {"Y", "y", "Yes", "yes"}
questionNo = {"N", "n", "No", "no"}

#Setup logging
user = getpass.getuser()

logger = logging.getLogger('Rubin Observatory XML Message Converter')

file_log_handler = logging.FileHandler('logfile.log')
logger.addHandler(file_log_handler)

stderr_log_handler = logging.StreamHandler()
logger.addHandler(stderr_log_handler)

class ContextFilter(logging.Filter):
    """
    This is a filter which injects contextual information into the log.
    """
    def filter(self, record):
        record.user = user
        return True

logger.addFilter(ContextFilter())

#Format logging output
formatter = logging.Formatter('%(asctime)s - %(name)s - %(user)s - %(levelname)s - %(message)s')
file_log_handler.setFormatter(formatter)
stderr_log_handler.setFormatter(formatter)

#Set the log level
logger.setLevel('DEBUG')

#Ask if Excel exports are required
#showExportQuestion(root)
needExport = needExport()
